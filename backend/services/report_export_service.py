import os
import csv
import io
from datetime import datetime
from decimal import Decimal
from typing import Dict, Any, List, Optional
from pathlib import Path

# PDF generation
from reportlab.lib.pagesizes import letter, A4, legal, landscape, portrait
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

# Excel generation
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import LineChart, Reference

import structlog
from schemas.report_schemas import ReportExportRequest, ReportFormat, FinancialReportData

logger = structlog.get_logger()

class ReportExportService:
    """Service for exporting reports to various formats"""
    
    def __init__(self):
        # Create exports directory if it doesn't exist
        self.export_dir = Path("/app/backend/exports")
        self.export_dir.mkdir(exist_ok=True)
    
    async def export_report(
        self,
        report_data: Dict[str, Any],
        export_request: ReportExportRequest,
        company_name: str,
        report_name: str
    ) -> Dict[str, Any]:
        """Export report to specified format"""
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        base_filename = f"{report_name.replace(' ', '_')}_{timestamp}"
        
        if export_request.format == ReportFormat.PDF:
            return await self._export_to_pdf(
                report_data, export_request, company_name, report_name, base_filename
            )
        elif export_request.format == ReportFormat.EXCEL:
            return await self._export_to_excel(
                report_data, export_request, company_name, report_name, base_filename
            )
        elif export_request.format == ReportFormat.CSV:
            return await self._export_to_csv(
                report_data, export_request, company_name, report_name, base_filename
            )
        else:
            raise ValueError(f"Unsupported export format: {export_request.format}")
    
    async def export_financial_report(
        self,
        financial_data: FinancialReportData,
        export_request: ReportExportRequest
    ) -> Dict[str, Any]:
        """Export financial report data"""
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        base_filename = f"{financial_data.report_name.replace(' ', '_')}_{timestamp}"
        
        if export_request.format == ReportFormat.PDF:
            return await self._export_financial_to_pdf(
                financial_data, export_request, base_filename
            )
        elif export_request.format == ReportFormat.EXCEL:
            return await self._export_financial_to_excel(
                financial_data, export_request, base_filename
            )
        elif export_request.format == ReportFormat.CSV:
            return await self._export_financial_to_csv(
                financial_data, export_request, base_filename
            )
        else:
            raise ValueError(f"Unsupported export format: {export_request.format}")
    
    async def _export_to_pdf(
        self,
        report_data: Dict[str, Any],
        export_request: ReportExportRequest,
        company_name: str,
        report_name: str,
        base_filename: str
    ) -> Dict[str, Any]:
        """Export report to PDF format"""
        
        filename = f"{base_filename}.pdf"
        filepath = self.export_dir / filename
        
        # Determine page size and orientation
        page_size = self._get_page_size(export_request.page_size)
        if export_request.page_orientation == "landscape":
            page_size = landscape(page_size)
        
        # Create PDF document
        doc = SimpleDocTemplate(
            str(filepath),
            pagesize=page_size,
            topMargin=0.75 * inch,
            bottomMargin=0.75 * inch,
            leftMargin=0.75 * inch,
            rightMargin=0.75 * inch
        )
        
        # Build content
        story = []
        styles = getSampleStyleSheet()
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Title'],
            fontSize=16,
            alignment=TA_CENTER,
            spaceAfter=12
        )
        
        story.append(Paragraph(report_name, title_style))
        story.append(Paragraph(company_name, styles['Normal']))
        story.append(Spacer(1, 12))
        
        # Add report date
        if 'generated_at' in report_data:
            date_str = report_data['generated_at'].strftime("%B %d, %Y at %I:%M %p")
            story.append(Paragraph(f"Generated: {date_str}", styles['Normal']))
            story.append(Spacer(1, 12))
        
        # Add data table
        if 'data' in report_data and report_data['data']:
            data_table = self._create_pdf_data_table(report_data['data'])
            story.append(data_table)
        
        # Add summary if requested
        if export_request.include_summary and 'summary' in report_data:
            story.append(Spacer(1, 12))
            story.append(Paragraph("Summary", styles['Heading2']))
            summary_table = self._create_pdf_summary_table(report_data['summary'])
            story.append(summary_table)
        
        # Build PDF
        doc.build(story)
        
        # Get file size
        file_size = filepath.stat().st_size
        
        logger.info("PDF report generated", filename=filename, file_size=file_size)
        
        return {
            "file_path": str(filepath),
            "filename": filename,
            "file_size": file_size,
            "format": ReportFormat.PDF
        }
    
    async def _export_financial_to_pdf(
        self,
        financial_data: FinancialReportData,
        export_request: ReportExportRequest,
        base_filename: str
    ) -> Dict[str, Any]:
        """Export financial report to PDF format"""
        
        filename = f"{base_filename}.pdf"
        filepath = self.export_dir / filename
        
        # Determine page size and orientation
        page_size = self._get_page_size(export_request.page_size)
        if export_request.page_orientation == "landscape":
            page_size = landscape(page_size)
        
        # Create PDF document
        doc = SimpleDocTemplate(
            str(filepath),
            pagesize=page_size,
            topMargin=0.75 * inch,
            bottomMargin=0.75 * inch,
            leftMargin=0.75 * inch,
            rightMargin=0.75 * inch
        )
        
        # Build content
        story = []
        styles = getSampleStyleSheet()
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Title'],
            fontSize=16,
            alignment=TA_CENTER,
            spaceAfter=12
        )
        
        story.append(Paragraph(financial_data.report_name, title_style))
        story.append(Paragraph(financial_data.company_name, styles['Normal']))
        
        # Report date
        date_str = financial_data.report_date.strftime("%B %d, %Y")
        story.append(Paragraph(f"As of {date_str}", styles['Normal']))
        
        if financial_data.comparison_date:
            comp_date_str = financial_data.comparison_date.strftime("%B %d, %Y")
            story.append(Paragraph(f"Compared to {comp_date_str}", styles['Normal']))
        
        story.append(Spacer(1, 20))
        
        # Add financial sections
        for section in financial_data.sections:
            story.append(Paragraph(section.section_name, styles['Heading2']))
            
            # Create section table
            section_table = self._create_financial_pdf_table(section, financial_data.comparison_date is not None)
            story.append(section_table)
            story.append(Spacer(1, 12))
        
        # Grand total if available
        if financial_data.grand_total is not None:
            story.append(Paragraph("Net Income" if "Profit" in financial_data.report_name else "Total Assets", styles['Heading2']))
            total_data = [[financial_data.grand_total]]
            total_table = Table(total_data)
            total_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, -1), colors.lightgrey),
                ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 12),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            story.append(total_table)
        
        # Build PDF
        doc.build(story)
        
        # Get file size
        file_size = filepath.stat().st_size
        
        logger.info("Financial PDF report generated", filename=filename, file_size=file_size)
        
        return {
            "file_path": str(filepath),
            "filename": filename,
            "file_size": file_size,
            "format": ReportFormat.PDF
        }
    
    async def _export_to_excel(
        self,
        report_data: Dict[str, Any],
        export_request: ReportExportRequest,
        company_name: str,
        report_name: str,
        base_filename: str
    ) -> Dict[str, Any]:
        """Export report to Excel format"""
        
        filename = f"{base_filename}.xlsx"
        filepath = self.export_dir / filename
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Report Data"
        
        # Header styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center")
        
        # Title
        ws.merge_cells('A1:F1')
        ws['A1'] = report_name
        ws['A1'].font = Font(bold=True, size=16)
        ws['A1'].alignment = Alignment(horizontal="center")
        
        # Company name
        ws.merge_cells('A2:F2')
        ws['A2'] = company_name
        ws['A2'].alignment = Alignment(horizontal="center")
        
        # Generated date
        if 'generated_at' in report_data:
            ws.merge_cells('A3:F3')
            ws['A3'] = f"Generated: {report_data['generated_at'].strftime('%B %d, %Y at %I:%M %p')}"
            ws['A3'].alignment = Alignment(horizontal="center")
        
        # Data section
        if 'data' in report_data and report_data['data']:
            data = report_data['data']
            if data:
                # Headers
                headers = list(data[0].keys())
                start_row = 5
                
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=start_row, column=col, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                
                # Data rows
                for row_idx, row_data in enumerate(data, start_row + 1):
                    for col, header in enumerate(headers, 1):
                        value = row_data.get(header, '')
                        if isinstance(value, Decimal):
                            value = float(value)
                        ws.cell(row=row_idx, column=col, value=value)
                
                # Freeze header row if requested
                if export_request.freeze_header_row:
                    ws.freeze_panes = f"A{start_row + 1}"
        
        # Summary sheet
        if export_request.include_summary and 'summary' in report_data:
            summary_ws = wb.create_sheet("Summary")
            summary_data = report_data['summary']
            
            row = 1
            for key, value in summary_data.items():
                summary_ws.cell(row=row, column=1, value=key).font = Font(bold=True)
                if isinstance(value, Decimal):
                    value = float(value)
                summary_ws.cell(row=row, column=2, value=value)
                row += 1
        
        # Auto-adjust column widths
        for ws in wb.worksheets:
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save workbook
        wb.save(str(filepath))
        
        # Get file size
        file_size = filepath.stat().st_size
        
        logger.info("Excel report generated", filename=filename, file_size=file_size)
        
        return {
            "file_path": str(filepath),
            "filename": filename,
            "file_size": file_size,
            "format": ReportFormat.EXCEL
        }
    
    async def _export_financial_to_excel(
        self,
        financial_data: FinancialReportData,
        export_request: ReportExportRequest,
        base_filename: str
    ) -> Dict[str, Any]:
        """Export financial report to Excel format"""
        
        filename = f"{base_filename}.xlsx"
        filepath = self.export_dir / filename
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = financial_data.report_name.replace(" ", "_")
        
        # Header styles
        title_font = Font(bold=True, size=16)
        header_font = Font(bold=True)
        currency_format = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
        
        # Title
        ws.merge_cells('A1:D1')
        ws['A1'] = financial_data.report_name
        ws['A1'].font = title_font
        ws['A1'].alignment = Alignment(horizontal="center")
        
        # Company name
        ws.merge_cells('A2:D2')
        ws['A2'] = financial_data.company_name
        ws['A2'].alignment = Alignment(horizontal="center")
        
        # Report date
        ws.merge_cells('A3:D3')
        date_str = financial_data.report_date.strftime("%B %d, %Y")
        ws['A3'] = f"As of {date_str}"
        ws['A3'].alignment = Alignment(horizontal="center")
        
        current_row = 5
        
        # Financial sections
        for section in financial_data.sections:
            # Section header
            ws.cell(row=current_row, column=1, value=section.section_name).font = header_font
            current_row += 1
            
            # Column headers
            if financial_data.comparison_date:
                ws.cell(row=current_row, column=1, value="Account").font = header_font
                ws.cell(row=current_row, column=2, value="Current").font = header_font
                ws.cell(row=current_row, column=3, value="Comparison").font = header_font
                ws.cell(row=current_row, column=4, value="Variance").font = header_font
            else:
                ws.cell(row=current_row, column=1, value="Account").font = header_font
                ws.cell(row=current_row, column=2, value="Amount").font = header_font
            current_row += 1
            
            # Section lines
            for line in section.lines:
                ws.cell(row=current_row, column=1, value=line.account_name)
                ws.cell(row=current_row, column=2, value=float(line.amount)).number_format = currency_format
                
                if financial_data.comparison_date and line.comparison_amount is not None:
                    ws.cell(row=current_row, column=3, value=float(line.comparison_amount)).number_format = currency_format
                    if line.variance_amount is not None:
                        ws.cell(row=current_row, column=4, value=float(line.variance_amount)).number_format = currency_format
                
                current_row += 1
            
            # Section total
            ws.cell(row=current_row, column=1, value=f"Total {section.section_name}").font = header_font
            ws.cell(row=current_row, column=2, value=float(section.total_amount)).font = header_font
            ws.cell(row=current_row, column=2).number_format = currency_format
            
            if financial_data.comparison_date and section.comparison_total is not None:
                ws.cell(row=current_row, column=3, value=float(section.comparison_total)).font = header_font
                ws.cell(row=current_row, column=3).number_format = currency_format
            
            current_row += 2
        
        # Grand total
        if financial_data.grand_total is not None:
            ws.cell(row=current_row, column=1, value="Net Income" if "Profit" in financial_data.report_name else "Total Assets").font = Font(bold=True, size=12)
            ws.cell(row=current_row, column=2, value=float(financial_data.grand_total)).font = Font(bold=True, size=12)
            ws.cell(row=current_row, column=2).number_format = currency_format
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save workbook
        wb.save(str(filepath))
        
        # Get file size
        file_size = filepath.stat().st_size
        
        logger.info("Financial Excel report generated", filename=filename, file_size=file_size)
        
        return {
            "file_path": str(filepath),
            "filename": filename,
            "file_size": file_size,
            "format": ReportFormat.EXCEL
        }
    
    async def _export_to_csv(
        self,
        report_data: Dict[str, Any],
        export_request: ReportExportRequest,
        company_name: str,
        report_name: str,
        base_filename: str
    ) -> Dict[str, Any]:
        """Export report to CSV format"""
        
        filename = f"{base_filename}.csv"
        filepath = self.export_dir / filename
        
        with open(filepath, 'w', newline='', encoding='utf-8') as csvfile:
            if 'data' in report_data and report_data['data']:
                data = report_data['data']
                if data:
                    # Get headers from first row
                    headers = list(data[0].keys())
                    
                    writer = csv.DictWriter(csvfile, fieldnames=headers)
                    
                    # Write header
                    csvfile.write(f"# {report_name}\n")
                    csvfile.write(f"# {company_name}\n")
                    if 'generated_at' in report_data:
                        csvfile.write(f"# Generated: {report_data['generated_at'].strftime('%B %d, %Y at %I:%M %p')}\n")
                    csvfile.write("\n")
                    
                    # Write data
                    writer.writeheader()
                    for row in data:
                        # Convert Decimal to string for CSV
                        csv_row = {}
                        for key, value in row.items():
                            if isinstance(value, Decimal):
                                csv_row[key] = str(value)
                            else:
                                csv_row[key] = value
                        writer.writerow(csv_row)
                    
                    # Write summary if requested
                    if export_request.include_summary and 'summary' in report_data:
                        csvfile.write("\n# Summary\n")
                        for key, value in report_data['summary'].items():
                            if isinstance(value, Decimal):
                                value = str(value)
                            csvfile.write(f"# {key}: {value}\n")
        
        # Get file size
        file_size = filepath.stat().st_size
        
        logger.info("CSV report generated", filename=filename, file_size=file_size)
        
        return {
            "file_path": str(filepath),
            "filename": filename,
            "file_size": file_size,
            "format": ReportFormat.CSV
        }
    
    async def _export_financial_to_csv(
        self,
        financial_data: FinancialReportData,
        export_request: ReportExportRequest,
        base_filename: str
    ) -> Dict[str, Any]:
        """Export financial report to CSV format"""
        
        filename = f"{base_filename}.csv"
        filepath = self.export_dir / filename
        
        with open(filepath, 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.writer(csvfile)
            
            # Header information
            writer.writerow([f"# {financial_data.report_name}"])
            writer.writerow([f"# {financial_data.company_name}"])
            writer.writerow([f"# As of {financial_data.report_date.strftime('%B %d, %Y')}"])
            if financial_data.comparison_date:
                writer.writerow([f"# Compared to {financial_data.comparison_date.strftime('%B %d, %Y')}"])
            writer.writerow([])
            
            # Data
            if financial_data.comparison_date:
                writer.writerow(["Section", "Account", "Current", "Comparison", "Variance"])
            else:
                writer.writerow(["Section", "Account", "Amount"])
            
            for section in financial_data.sections:
                for line in section.lines:
                    if financial_data.comparison_date:
                        writer.writerow([
                            section.section_name,
                            line.account_name,
                            str(line.amount),
                            str(line.comparison_amount) if line.comparison_amount else "",
                            str(line.variance_amount) if line.variance_amount else ""
                        ])
                    else:
                        writer.writerow([
                            section.section_name,
                            line.account_name,
                            str(line.amount)
                        ])
                
                # Section total
                if financial_data.comparison_date:
                    writer.writerow([
                        section.section_name,
                        f"Total {section.section_name}",
                        str(section.total_amount),
                        str(section.comparison_total) if section.comparison_total else "",
                        ""
                    ])
                else:
                    writer.writerow([
                        section.section_name,
                        f"Total {section.section_name}",
                        str(section.total_amount)
                    ])
                
                writer.writerow([])  # Empty row between sections
            
            # Grand total
            if financial_data.grand_total is not None:
                total_label = "Net Income" if "Profit" in financial_data.report_name else "Total Assets"
                writer.writerow(["", total_label, str(financial_data.grand_total)])
        
        # Get file size
        file_size = filepath.stat().st_size
        
        logger.info("Financial CSV report generated", filename=filename, file_size=file_size)
        
        return {
            "file_path": str(filepath),
            "filename": filename,
            "file_size": file_size,
            "format": ReportFormat.CSV
        }
    
    def _get_page_size(self, page_size_name: str):
        """Get page size object from name"""
        sizes = {
            "letter": letter,
            "a4": A4,
            "legal": legal
        }
        return sizes.get(page_size_name.lower(), letter)
    
    def _create_pdf_data_table(self, data: List[Dict[str, Any]]) -> Table:
        """Create PDF table from data"""
        if not data:
            return Table([["No data available"]])
        
        # Headers
        headers = list(data[0].keys())
        table_data = [headers]
        
        # Data rows
        for row in data:
            table_row = []
            for header in headers:
                value = row.get(header, '')
                if isinstance(value, Decimal):
                    value = f"${value:,.2f}"
                table_row.append(str(value))
            table_data.append(table_row)
        
        # Create table
        table = Table(table_data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        return table
    
    def _create_pdf_summary_table(self, summary: Dict[str, Any]) -> Table:
        """Create PDF summary table"""
        table_data = []
        for key, value in summary.items():
            if isinstance(value, Decimal):
                value = f"${value:,.2f}"
            table_data.append([key, str(value)])
        
        table = Table(table_data)
        table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        return table
    
    def _create_financial_pdf_table(self, section, has_comparison: bool) -> Table:
        """Create PDF table for financial section"""
        if has_comparison:
            table_data = [["Account", "Current", "Comparison", "Variance"]]
            for line in section.lines:
                table_data.append([
                    line.account_name,
                    f"${line.amount:,.2f}",
                    f"${line.comparison_amount:,.2f}" if line.comparison_amount else "",
                    f"${line.variance_amount:,.2f}" if line.variance_amount else ""
                ])
            table_data.append([
                f"Total {section.section_name}",
                f"${section.total_amount:,.2f}",
                f"${section.comparison_total:,.2f}" if section.comparison_total else "",
                ""
            ])
        else:
            table_data = [["Account", "Amount"]]
            for line in section.lines:
                table_data.append([line.account_name, f"${line.amount:,.2f}"])
            table_data.append([f"Total {section.section_name}", f"${section.total_amount:,.2f}"])
        
        table = Table(table_data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        return table